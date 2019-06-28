def process(Report_add,GDW_add,GOC_add):
    print('started')
    import sys
    import numpy as np
    import datetime
    import os
    import openpyxl
    from openpyxl.styles.colors import Color
    from openpyxl.styles.colors import WHITE
    from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
    reload(sys)
    sys.setdefaultencoding('utf8')
    import pandas as pd


    pd.read_excel(Report_add, sheet_name='Data').to_csv('main.csv', index=False)
    main = pd.read_csv("main.csv")
    os.remove("main.csv")

    pd.read_excel(GDW_add).to_csv('GDW.csv', index=False)
    gdw = pd.read_csv("GDW.csv")
    os.remove("GDW.csv")

    pd.read_excel(GOC_add).to_csv('GOC.csv', index=False)
    goc = pd.read_csv("GOC.csv")
    os.remove("GOC.csv")
    
    #Removing unwanted top rows
    main =main.iloc[15:]

    #Function that finds effective working days in each week
    def Working_days(start_date,date_string):
        
        if(type(start_date) != unicode):
            start_date = u'2000-10-27 00:00:00'
    
            
        
        arr = date_string.split()
        arr.remove(u'Week')
        

        brr = arr[0].split('-')

        start = brr[0]
        end = brr[1]
        d1 = start.split('/')
        d2 = end.split('/')

        month = int(d2[0])
        day = int(d2[1])
        year = int(d2[2])


        if month == 1 and day == 31 :
            month = 2
            day = 1

        if month == 2 and day == 29 :
            month = 3
            day = 1    

        if month == 2 and day == 28 and year%4==0 :
            month = 2
            day = 29    

        if month == 3 and day == 31 :
            month = 4
            day = 1       

        if month == 5 and day == 31 :
            month = 6
            day = 1     

        if month == 7 and day == 31 :
            month=8
            day = 1     

        if month == 8 and day == 31 :
            month=9
            day = 1     

        if month == 10 and day == 31 :
            month=11
            day = 1     

        if month == 12 and day == 31 :
            month = 1
            day = 1 
            year=year +1


        if month == 4 and day == 30 :
            month = 5
            day = 1     

        if month == 6 and day == 30 :
            month = 7
            day = 1       


        if month == 9 and day == 30 :
            month = 10
            day = 1       


        if month == 11 and day == 30 :
            month = 12
            day = 1       


        d2[0] = str(month)
        d2[1] = str(day)
        d2[2] = str(year)

        if(len(d1[0])==1):
            d1[0] = u'0' + d1[0]
        if(len(d1[1])==1):
            d1[1] = u'0' + d1[1]
        if(len(d2[0])==1):
            d2[0] = u'0' + d2[0]
        if(len(d2[1])==1):
            d2[1] = u'0' + d2[1]    

        begin = d1[2] + '-' + d1[0] + '-' + d1[1] 
        finish = d2[2] + '-' + d2[0] + '-' + d2[1]  
    
        
        
        a= start_date.split()
        a.remove(a[1])

        a = a[0].split('-')

        start_month = a[1]
        start_day = a[2]
        start_year = a[0]
        
        if(len(a[0])==1):
            a[0] = u'0' + a[0]
        if(len(a[1])==1):
            a[1] = u'0' + a[1]
        
        Rstart = a[0] + '-' + a[1] + '-' + a[2]
  
        count1 = np.busday_count(Rstart,begin)
        count2 = np.busday_count(Rstart,finish)
        
        if(count1 >=0):
            return np.busday_count(begin,finish)
        else:
            return max(0,np.busday_count(Rstart,finish))
        
        
        
    

    #Removing unwanted columns
    main = main.drop(columns = "TRS Monthly Report")
    main = main.drop(columns = "Unnamed: 3")
    # main = main.drop(columns = "Unnamed: 4")
    main = main.drop(columns = "Unnamed: 8")
    main = main.drop(columns = "Unnamed: 9")
    main = main.drop(columns = "Unnamed: 10")
    main = main.drop(columns = "Unnamed: 12")
    main = main.drop(columns = "Unnamed: 14")
    main = main.drop(columns = "Unnamed: 20")
    main = main.drop(columns = "Unnamed: 22")
    main = main.drop(columns = "Unnamed: 16")
    main = main.drop(columns = "Unnamed: 18")
    main = main.drop(columns = "Unnamed: 24")
    main = main.drop(columns = "Unnamed: 31")
    # main = main.drop(columns = "Unnamed: 32")
    # main = main.drop(columns = "Unnamed: 33")

     #Using first row as name of columns
    main.columns = main.loc[15]
    main = main.iloc[1:]


   
    #if resource code column has LOA delete that row
    #if max expected hours column has 0 delete that row


    

    #Removing LOA
    indexNames = main[main['Resource Code']=='LOA'].index
    main.drop(indexNames ,inplace = True)

    #Removing max expected hours = 0
    indexNames2 = main[main['Monthly Expected Hours']=='0'].index
    main.drop(indexNames2 ,inplace = True)
   
    #Dropping Resource code column after use
    main = main.drop(columns = "Resource Code")

    #merging goc file
    result = pd.merge(goc,main,
                    on='System GOC')

    result = result.drop(columns = "SMT Direct")
    

    #matching the case of resource soeid column
    gdw['Resource SOEID'] = gdw['Resource SOEID'].str.lower()
   
    #Merging gdw with result
    result2 = pd.merge(result,gdw[['Resource SOEID', 'Location','LOB','Region']],on ='Resource SOEID')
   
    LOB = result2['LOB']
    result3 = result2.drop(columns = "LOB")
    SystemGOC = result2['System GOC']
    result3 = result3.drop(columns = "System GOC")


   

    #Rearranging the columns
    result3.insert(loc = 0, column = 'AMH LOB SubArea',value = LOB)
    result3.insert(loc = 7, column = 'System GOC',value = SystemGOC)




    result3.to_excel(r'C:\Users\ak47354\Desktop\TRS Automation\temp.xlsx',index = None , header = True,encoding = 'utf-8')
    os.remove("temp.xlsx")

    final =result3.sort_values(by=['AMH LOB SubArea','Direct','Resource Name'])

    final['Resource SOEID'] = final['Resource SOEID'].str.upper()
    final['Resource Manager SOEID'] = final['Resource Manager SOEID'].str.upper()
   
    now = datetime.datetime.now()
    output_name = 'TRS Monthly Report as on ' + str(now.day) + '-'+ str(now.month) + '-'+ str(now.year)
    output_name += '.xlsx'
   
    final.to_excel(output_name)

        
    
    wb = openpyxl.load_workbook(output_name)
    type(wb)
    os.remove(output_name)
    sheet = wb.active
    row_count = sheet.max_row
    column_count = sheet.max_column
    
    factor = (1/7.0)
    # 1 unit this equals 7 pixels in excel
    sheet.column_dimensions['A'].width = 0
    sheet.column_dimensions['B'].width = factor*73
    sheet.column_dimensions['C'].width = factor*107
    sheet.column_dimensions['D'].width = factor*198
    sheet.column_dimensions['E'].width = factor*62
    sheet.column_dimensions['F'].width = factor*160
    sheet.column_dimensions['G'].width = factor*96
    sheet.column_dimensions['H'].width = factor*86
    sheet.column_dimensions['I'].width = factor*60
    sheet.column_dimensions['J'].width = factor*144
    sheet.column_dimensions['K'].width = factor*144
    sheet.column_dimensions['L'].width = factor*144
    sheet.column_dimensions['M'].width = factor*144
    sheet.column_dimensions['N'].width = factor*144
    sheet.column_dimensions['O'].width = factor*144
    sheet.column_dimensions['P'].width = factor*144
    sheet.column_dimensions['Q'].width = factor*144
    sheet.column_dimensions['R'].width = factor*144
    sheet.column_dimensions['S'].width = factor*144
    sheet.column_dimensions['T'].width = factor*133
    sheet.column_dimensions['U'].width = factor*71
    sheet.column_dimensions['V'].width = factor*71


    

    for x in range(11,21):
        for y in range(2,row_count + 1):
            sheet.cell(row=y,column = x).value = float(sheet.cell(row=y,column = x).value)

    
    SZ14_FONT = Font(name='Arial',
                    size=8,
                    bold=False,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='000000')

    for row in sheet:
        for cell in row:
            cell.font = SZ14_FONT

    w1 = sheet.cell(row=1,column = 11).value
    w2 = sheet.cell(row=1,column = 12).value
    w3 = sheet.cell(row=1,column = 13).value
    w4 = sheet.cell(row=1,column = 14).value
    w5 = sheet.cell(row=1,column = 15).value

    blueFill = PatternFill(start_color='0000FF',
                    end_color='0000FF',
                    fill_type='solid',
                        )

    redFill = PatternFill(start_color='FFFF0000',
                    end_color='FFFF0000',
                    fill_type='solid')

    for x in range (1,column_count+1):
        sheet.cell(row=1,column = x).fill = blueFill
        sheet.cell(row=1,column = x).font = ft = Font(color=WHITE)
        



    for x in range (2,row_count+1):
        target = sheet.cell(row=x,column = 11).value
        target = float(target)
        start_date = sheet.cell(row=x,column = 10).value

        wd = Working_days(start_date,w1)
        wd = float(wd)
        mwh = sheet.cell(row=x,column = 19).value
        mwh = float(mwh)
        if(target < ((wd/5.0)*mwh)):
            sheet.cell(row=x,column = 11).fill = redFill
            if  sheet.cell(row=x,column = 21).value != 'All Time to Non-Project' :
                sheet.cell(row=x,column = 21).value = 'Less Than Expected'

    for x in range (2,row_count+1):
        target = sheet.cell(row=x,column = 12).value
        target = float(target)
        start_date = sheet.cell(row=x,column = 10).value
        wd = Working_days(start_date,w2)
        wd = float(wd)
        mwh = sheet.cell(row=x,column = 19).value
        mwh = float(mwh)
        if(target < ((wd/5.0)*mwh)):
            sheet.cell(row=x,column = 12).fill = redFill
            if  sheet.cell(row=x,column = 21).value != 'All Time to Non-Project' :
                sheet.cell(row=x,column = 21).value = 'Less Than Expected'
            
    for x in range (2,row_count+1):
        target = sheet.cell(row=x,column = 13).value
        target = float(target)
        start_date = sheet.cell(row=x,column = 10).value
    
        wd = Working_days(start_date,w3)
        wd = float(wd)
        mwh = sheet.cell(row=x,column = 19).value
        mwh = float(mwh)
        if(target < ((wd/5.0)*mwh)):
            sheet.cell(row=x,column = 13).fill = redFill
            if  sheet.cell(row=x,column = 21).value != 'All Time to Non-Project' :
                sheet.cell(row=x,column = 21).value = 'Less Than Expected'
            
    for x in range (2,row_count+1):
        target = sheet.cell(row=x,column = 14).value
        target = float(target)
        start_date = sheet.cell(row=x,column = 10).value
        wd = Working_days(start_date,w4)
        wd = float(wd)
        mwh = sheet.cell(row=x,column = 19).value
        mwh = float(mwh)
        if(target < ((wd/5.0)*mwh)):
            sheet.cell(row=x,column = 14).fill = redFill
            if  sheet.cell(row=x,column = 21).value != 'All Time to Non-Project' :
                sheet.cell(row=x,column = 21).value = 'Less Than Expected'
            
    for x in range (2,row_count+1):
        target = sheet.cell(row=x,column = 15).value
        target = float(target)
        start_date = sheet.cell(row=x,column = 10).value
        wd = Working_days(start_date,w5)
        wd = float(wd)
        mwh = sheet.cell(row=x,column = 19).value
        mwh = float(mwh)
        totalhrs = sheet.cell(row=x,column = 18).value
        if(target < ((wd/5.0)*mwh)):
            sheet.cell(row=x,column = 15).fill = redFill
            if  sheet.cell(row=x,column = 21).value != 'All Time to Non-Project' :
                sheet.cell(row=x,column = 21).value = 'Less Than Expected'
                
    for x in range (2,row_count+1): 
        #correct this for the resource start date case
            if  sheet.cell(row=x,column = 18).value == 0 :
                sheet.cell(row=x,column = 21).value = 'No time Booked'

    for x in range (2,row_count+1):
        value =sheet.cell(row=x,column = 21).value
        if  value != 'All Time to Non-Project' and value != 'Less Than Expected' and value !='No time Booked':
            sheet.cell(row=x,column = 21).value = 'Completed'            
                
                
    sheet.delete_cols(10)
    sheet.delete_cols(21)
    sheet.delete_cols(21)
                
    wb.save(output_name) 
    #this overwrites existing files with same name
    from shutil import move
    move( str(output_name), "Processed_Reports/"+ str(output_name))

    print('completed')
